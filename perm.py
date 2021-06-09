from openpyxl import Workbook
from openpyxl import load_workbook
from fuzzywuzzy import fuzz




def urez_perm(a, sheet, row):
    if 'рабочий поселок' in a:
        a = a.replace('рабочий поселок', 'рп')
    if ' г ' in a:
        a = a.replace(' г ', '')
    if ' г,' in a:
        a = a.replace(' г,', '')
    if 'г. ' in a:
        a = a.replace('г. ', '')
    if 'станция метро' in a:
        a = a.replace('станция метро', 'ст. метро')
    if 'шоссе' in a:
        a = a.replace('шоссе', 'ш.')
    if 'пос.' in a:
        a = a.replace('пос.', 'п.')
    if 'район' in a:
        a = a.replace('район', 'рн.')
    if 'р-он' in a:
        a = a.replace('р-он', 'рн.')
    if 'р-н' in a:
        a = a.replace('р-н', 'рн.')
    if 'улица' in a:
        a = a.replace('улица', 'ул.')
    if 'проспект' in a:
        a = a.replace('проспект', 'пр.')
    if 'пр-кт' in a:
        a = a.replace('пр-кт', 'пр.')
    if 'пр-т' in a:
        a = a.replace('пр-т', 'пр.')
    if 'пос.' in a:
        a = a.replace('пос.', 'п.')
    if 'дом' in a:
        a = a.replace('дом', 'д.')
    if 'собственный' in a:
        a = a.replace('собственный', '')
    if 'собственная' in a:
        a = a.replace('собственная', '')
    if 'столб' in a:
        a = a.replace('столб', 'АМС')
    if 'Столб' in a:
        a = a.replace('Столб', 'АМС')
    if 'опора' in a:
        a = a.replace('опора', 'АМС')
    if 'Опора' in a:
        a = a.replace('Опора', 'АМС')
    if 'башня' in a:
        a = a.replace('башня', 'АМС')
    if 'Башня' in a:
        a = a.replace('Башня', 'АМС')
    if 'вышка' in a:
        a = a.replace('вышка', 'АМС')
    if 'Вышка' in a:
        a = a.replace('Вышка', 'АМС')
    if 'мачта' in a:
        a = a.replace('мачта', 'АМС')
    if 'Мачта' in a:
        a = a.replace('Мачта', 'АМС')
    if 'ОАО' in a:
        a = a.replace('ОАО', '')
    if 'ПАО' in a:
        a = a.replace('ПАО', '')
    if 'ООО' in a:
        a = a.replace('ООО', '')
    if '«' in a:
        a = a.replace('«', '"')
    if '»' in a:
        a = a.replace('»', '"')
    if 'МегаФон' in a:
        a = a.replace('МегаФон', 'МФ')
    if 'ВымпелКом' in a:
        a = a.replace('ВымпелКом', 'ВК')
    if 'Вымпел-Ком' in a:
        a = a.replace('Вымпел-Ком', 'ВК')
    if 'Вымпел-Коммуникации' in a:
        a = a.replace('Вымпел-Коммуникации', 'ВК')
    if 'Т2 РТК Холдинг' in a:
        a = a.replace('Т2 РТК Холдинг', 'Т2')
    if 'Т2 Мобайл' in a:
        a = a.replace('Т2 Мобайл', 'Т2')
    if 'Теле2' in a:
        a = a.replace('Теле2', 'Т2')
    if 'ТЕЛЕ2' in a:
        a = a.replace('ТЕЛЕ2', 'Т2')
    if '"Теле2-Н.Новгород"' in a:
        a = a.replace('"Теле2-Н.Новгород"', 'Т2')
    if 'Дом' in a:
        a = a.replace('Дом', 'д.')
    sheet.cell(row=row, column=9).value = a

def poisk_perm(technologia, lst):
    if technologia in wsSnat.cell(row=row11, column=6).value:
        e=wsSnat.cell(row=row11, column=9).value
        b = 0
        for i in lst:
            c=fuzz.WRatio(e,i)
            if c>b:
                b=c
        if b<87:
            li = []
            for col in range(2, wsSnat.max_column - 1):
                li.append(wsSnat.cell(row=row11, column=col).value)
            wsResult.append(li)


def proga_perm(zaregi, istekli, katalog):
    global wsSearch
    wbSearch = load_workbook(zaregi)
    wsSearch = wbSearch.active
    global wsSnat
    wbSnat = load_workbook(istekli)
    wsSnat = wbSnat.active
    global wsResult
    wbResult = Workbook()
    wsResult = wbResult.active
    mf_adresgsm = []
    mf_adresumts = []
    mf_adreslte = []
    mf_adresrrl = []
    mts_adresgsm = []
    mts_adresumts = []
    mts_adreslte = []
    mts_adresrrl = []
    vk_adresgsm = []
    vk_adresumts = []
    vk_adreslte = []
    vk_adresrrl = []
    t2_adresgsm = []
    t2_adresumts = []
    t2_adreslte = []
    t2_adresrrl = []
    mf_adres = []
    mts_adres = []
    vk_adres = []
    t2_adres = []
    other_adres = []

    for row in range(9, wsSearch.max_row-5):
        urez_perm(wsSearch.cell(row=row, column=9).value, wsSearch, row)
        if 'Мега' in wsSearch.cell(row=row, column=4).value:
            if '18.1.1.3' in wsSearch.cell(row=row, column=6).value:
                mf_adresgsm.append(wsSearch.cell(row=row, column=9).value)
            elif '18.1.1.5' in wsSearch.cell(row=row, column=6).value:
                mf_adresumts.append(wsSearch.cell(row=row, column=9).value)
            elif '18.7.1' in wsSearch.cell(row=row, column=6).value:
                mf_adreslte.append(wsSearch.cell(row=row, column=9).value)
            elif '19.2' in wsSearch.cell(row=row, column=6).value:
                mf_adresrrl.append(wsSearch.cell(row=row, column=9).value)
            else: mf_adres.append(wsSearch.cell(row=row, column=9).value)

        elif 'Вымпел' in wsSearch.cell(row=row, column=4).value:
            if '18.1.1.3' in wsSearch.cell(row=row, column=6).value:
                vk_adresgsm.append(wsSearch.cell(row=row, column=9).value)
            elif '18.1.1.5' in wsSearch.cell(row=row, column=6).value:
                vk_adresumts.append(wsSearch.cell(row=row, column=9).value)
            elif '18.7.1' in wsSearch.cell(row=row, column=6).value:
                vk_adreslte.append(wsSearch.cell(row=row, column=9).value)
            elif '19.2' in wsSearch.cell(row=row, column=6).value:
                vk_adresrrl.append(wsSearch.cell(row=row, column=9).value)
            else: vk_adres.append(wsSearch.cell(row=row, column=9).value)

        elif 'Мобильные' in wsSearch.cell(row=row, column=4).value:
            if '18.1.1.3' in wsSearch.cell(row=row, column=6).value:
                mts_adresgsm.append(wsSearch.cell(row=row, column=9).value)
            elif '18.1.1.5' in wsSearch.cell(row=row, column=6).value:
                mts_adresumts.append(wsSearch.cell(row=row, column=9).value)
            elif '18.7.1' in wsSearch.cell(row=row, column=6).value:
                mts_adreslte.append(wsSearch.cell(row=row, column=9).value)
            elif '19.2' in wsSearch.cell(row=row, column=6).value:
                mts_adresrrl.append(wsSearch.cell(row=row, column=9).value)
            else: mts_adres.append(wsSearch.cell(row=row, column=9).value)

        elif 'Т2 Мобайл' in wsSearch.cell(row=row, column=4).value:
            if '18.1.1.3' in wsSearch.cell(row=row, column=6).value:
                t2_adresgsm.append(wsSearch.cell(row=row, column=9).value)
            elif '18.1.1.5' in wsSearch.cell(row=row, column=6).value:
                t2_adresumts.append(wsSearch.cell(row=row, column=9).value)
            elif '18.7.1' in wsSearch.cell(row=row, column=6).value:
                t2_adreslte.append(wsSearch.cell(row=row, column=9).value)
            elif '19.2' in wsSearch.cell(row=row, column=6).value:
                t2_adresrrl.append(wsSearch.cell(row=row, column=9).value)
            else: t2_adres.append(wsSearch.cell(row=row, column=9).value)
        elif ('6.1.1.' not in wsSearch.cell(row=row, column=6).value
              and '41.2' not in wsSearch.cell(row=row, column=6).value
              and '41.7.' not in wsSearch.cell(row=row, column=6).value
              and '18.1.9.' not in wsSearch.cell(row=row, column=6).value
              and '18.2.3.' not in wsSearch.cell(row=row, column=6).value
              and '18.2.1.' not in wsSearch.cell(row=row, column=6).value
              and '18.2.6.' not in wsSearch.cell(row=row, column=6).value
              and '18.2.8.' not in wsSearch.cell(row=row, column=6).value
              and '19.4.4.2.' not in wsSearch.cell(row=row, column=6).value):
                other_adres.append(wsSearch.cell(row=row, column=9).value)

    firstrow = []
    for col in range(2, wsSnat.max_column - 1):
        firstrow.append(wsSnat.cell(row=7, column=col).value)
    wsResult.append(firstrow)

    for row1 in range(9, wsSnat.max_row-5):
            global row11
            row11=row1
            urez_perm(wsSnat.cell(row=row1, column=9).value, wsSnat, row1)

            if 'Мега' in wsSnat.cell(row=row1, column=4).value:
                poisk_perm('18.1.1.3', mf_adresgsm)
                poisk_perm('18.1.1.5', mf_adresumts)
                poisk_perm('18.7.1', mf_adreslte)
                poisk_perm('19.2', mf_adresrrl)
                if '18.1.1.3' not in wsSnat.cell(row=row1, column=6).value and '18.1.1.5' not in wsSnat.cell(row=row1, column=6).value and '18.7.1.' not in wsSnat.cell(row=row1, column=6).value and '19.2' not in wsSnat.cell(row=row1, column=6).value:
                    if wsSnat.cell(row=row1, column=9).value not in mf_adres:
                        li = []
                        for col in range(2, wsSnat.max_column - 1):
                            li.append(wsSnat.cell(row=row1, column=col).value)
                        wsResult.append(li)
            elif 'Вымпел' in wsSnat.cell(row=row1, column=3).value:
                poisk_perm('18.1.1.3', vk_adresgsm)
                poisk_perm('18.1.1.5', vk_adresumts)
                poisk_perm('18.7.1', vk_adreslte)
                poisk_perm('19.2', vk_adresrrl)
                if '18.1.1.3' not in wsSnat.cell(row=row1, column=6).value and '18.1.1.5' not in wsSnat.cell(row=row1, column=6).value and '18.7.1.' not in wsSnat.cell(row=row1, column=6).value and '19.2' not in wsSnat.cell(row=row1, column=6).value:
                    if wsSnat.cell(row=row1, column=9).value not in vk_adres:
                        li = []
                        for col in range(2, wsSnat.max_column - 1):
                            li.append(wsSnat.cell(row=row1, column=col).value)
                        wsResult.append(li)
            elif 'Мобильные' in wsSnat.cell(row=row1, column=3).value:
                poisk_perm('18.1.1.3', mts_adresgsm)
                poisk_perm('18.1.1.5', mts_adresumts)
                poisk_perm('18.7.1', mts_adreslte)
                poisk_perm('19.2', mts_adresrrl)
                if '18.1.1.3' not in wsSnat.cell(row=row1, column=6).value and '18.1.1.5' not in wsSnat.cell(row=row1, column=6).value and '18.7.1.' not in wsSnat.cell(row=row1, column=6).value and '19.2' not in wsSnat.cell(row=row1, column=6).value:
                    if wsSnat.cell(row=row1, column=9).value not in mts_adres:
                        li = []
                        for col in range(2, wsSnat.max_column - 1):
                            li.append(wsSnat.cell(row=row1, column=col).value)
                        wsResult.append(li)
            elif 'Т2 Мобайл' in wsSnat.cell(row=row1, column=4).value:
                poisk_perm('18.1.1.3', t2_adresgsm)
                poisk_perm('18.1.1.5', t2_adresumts)
                poisk_perm('18.7.1', t2_adreslte)
                poisk_perm('19.2', t2_adresrrl)
                if '18.1.1.3' not in wsSnat.cell(row=row1, column=6).value and '18.1.1.5' not in wsSnat.cell(row=row1, column=6).value and '18.7.1.' not in wsSnat.cell(row=row1, column=6).value and '19.2' not in wsSnat.cell(row=row1, column=6).value:
                    if wsSnat.cell(row=row1, column=9).value not in t2_adres:
                        li = []
                        for col in range(2, wsSnat.max_column - 1):
                            li.append(wsSnat.cell(row=row1, column=col).value)
                        wsResult.append(li)
            elif ('6.1.1.' not in wsSnat.cell(row=row1, column=6).value
              and '41.2' not in wsSnat.cell(row=row1, column=6).value
              and '41.7.' not in wsSnat.cell(row=row1, column=6).value
              and '18.1.9.' not in wsSnat.cell(row=row1, column=6).value
              and '18.2.3.' not in wsSnat.cell(row=row1, column=6).value
              and '18.2.1.' not in wsSnat.cell(row=row1, column=6).value
              and '18.2.6.' not in wsSnat.cell(row=row1, column=6).value
              and '18.2.8.' not in wsSnat.cell(row=row1, column=6).value
              and '19.4.4.2.' not in wsSnat.cell(row=row1, column=6).value):
                    if wsSnat.cell(row=row1, column=9).value not in other_adres:
                        li = []
                        for col in range(2, wsSnat.max_column - 1):
                            li.append(wsSnat.cell(row=row1, column=col).value)
                        wsResult.append(li)

    wbResult.save(f'{katalog}/Итоговый список незарегистрированных РЭС.xlsx')

